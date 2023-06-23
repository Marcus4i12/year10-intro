import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.active

worksheet.cell(row=1, column=1).value = "Name:"
worksheet.cell(row=1, column=2).value = "Age:"
worksheet.cell(row=1, column=3).value = "Networth:"

while True:
    name = input("What's your name? (type 'end' to stop) ")
    if name == "end":
        break
    age = input("What's your age? ")
    networth = input("What's your net worth? ")
    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1).value = name
    worksheet.cell(row=row, column=2).value = age
    worksheet.cell(row=row, column=3).value = networth

workbook.save(r"c:\Users\training.user\Downloads\flash.csv")